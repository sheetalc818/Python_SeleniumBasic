import openpyxl

path = "/home/sheetal/PycharmProjects/Selenium_Basics/Resources/ExcelFile/testing.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
print("Number of Rows", rows)
print("Number of Column", cols)
print()
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(row=r, column=c).value, end="  ")
    print()

# To Write data into excel file
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        sheet.cell(row=r, column=c).value = "Hello"

workbook.save(path)
